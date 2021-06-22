'''
读取excle文件中测试数据
读取yaml中的数据
写入数据到yaml中
'''
from openpyxl import load_workbook
from openpyxl import Workbook
import yaml
import Log

log = Log.Log()

class ReadData(object):

    def read_excel(self,row,col,path,sheet_name):  # 获取数据函数
        #self.wb = load_workbook('C:\\Programs\\PycharmProjects\\ShangMa_Admin\\testData\\test_data.xlsx',read_only=True)
        try:
            self.wb = load_workbook(path)
            self.workSheet = self.wb[sheet_name]  # 获取当前表
            self.data = self.workSheet.cell(row,col).value
            self.wb.close()
            return self.data
        except Exception as e:
            log.error("read_excle类出现错误%s"%e)
            self.wb.close()
            return None


    #清空excle表中特定列的数据
    def del_col(self,path,sheet_name,col):
        try:
            self.wb = load_workbook(path)
            self.workSheet = self.wb[sheet_name]
            self.workSheet.delete_cols(col)
            self.workSheet.save(path)
        except Exception as e:
            log.error("del_excle类出现错误%s"%e)

    # 删除行数据
    def del_row(self, path, sheet_name, row):
        try:
            self.wb = load_workbook(path)
            self.workSheet = self.wb[sheet_name]
            self.workSheet.delete_rows(row)
            self.workSheet.save(path)
        except Exception as e:
            log.error("del_excle类出现错误%s" % e)

    #获取工作表的有效行数
    def get_max_row(self,path,sheet_name):
        self.wb = load_workbook(path)
        self.workSheet = self.wb[sheet_name]  # 获取当前表
        self.endRow = self.workSheet.max_row  # 获取有效工作行数
        return self.endRow


    #读取yaml文件
    def read_yaml(self,key):
        self.f = open('projects_utils/api_config.yml', encoding='utf-8')
        self.data = yaml.load(self.f.read(),Loader=yaml.Loader) #读取yaml，读取后data的格式是字典；如果直接用f.read()读取，读取的格式是str
        self.f.close()
        return self.data[key]



    #写入数据到yaml中
    def write_yaml(self):
        pass

    def write_excle(self,data,path,sheet_name,row,column):
        # self.wb = Workbook(write_only=True)
        # ws = self.wb.create_sheet(sheet_name)
        try:
            self.wb = load_workbook(path)
            self.workSheet = self.wb[sheet_name]
            _=self.workSheet.cell(row =row,column =column,value =str(data))
            self.wb.save(path)
            self.wb.close()
        except Exception as e :
            log.error(e)


if __name__== '__main__':
    ya = ReadData()
    # 测试用例的相对路径
    path = "..\\testData\\test_data.xlsx"
    data = ya.read_excel(2,2,path,'admin_test')
    print(data)















